import openpyxl
import logging
#####
import logging_factory

#####
logger = logging_factory.get_module_logger("excel_reader", logging.DEBUG)
logger_err = logging_factory.get_module_logger("excel_reader_err", logging.ERROR)


def get_queries_and_scales(path: str, start_row: int, scales_column: int, queries_column: int, related_column: int):
    """
    Function to obtain the scales' names and queries

    :param path: str - the path where the excel spreadsheet is located
    :param start_row: int - row number where the first query is located
    :param scales_column: int - column number where the scales are located
    :param queries_column: int - column number where the queries are located
    :param related_column: int - column number where the relationship between queries in different scales is located
    :return: result: dict - a dictionary with scales as keys and list of queries as values

    """

    # Workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row

    result = {}
    current_scale = ""

    for i in range(start_row, m_row + 1):
        scale = sheet_obj.cell(row=i, column=scales_column).value

        # A new scale appears
        if scale not in result and scale is not None:
            current_scale = scale
            result[scale] = []

        # For merged cells
        if scale is None:
            scale = current_scale

        query = sheet_obj.cell(row=i, column=queries_column).value
        queries = []

        related_code = sheet_obj.cell(row=i, column=related_column).value
        codes = []
        related_scales = []

        if related_code is not None:
            related_code = str(related_code)  # Cast to string just in case we have to split it later
            to_check_codes = convert_code(related_code)
            for c in to_check_codes:
                if code_present(wb_obj, c, start_row, related_column):
                    codes.append(c)
                else:
                    logger_err.error("Code '{}' appears only one time".format(c))
            related_scales = get_related_scales(wb_obj, current_scale, codes, start_row, scales_column, related_column)

        to_add = []
        if query is not None:
            if "," in query:  # Multiple queries in the same cell
                queries = query.split(",")
            if len(queries) > 0:
                for q in queries:
                    to_add.append(q.strip())
                    to_add.append(codes)
                    to_add.append(related_scales)
                    result[scale].append(to_add)
                    to_add = []
            # Single query
            else:
                to_add.append(query.strip())
                to_add.append(codes)
                to_add.append(related_scales)
                result[scale].append(to_add)

    logger.debug("Scales and queries loaded")

    return result


def get_related_scales(wb_obj, current_scale: str, codes: list, start_row: int, scales_column: int, related_column: int):
    """
    Function that returns the name of all the scales related to the relationship code given

    :param wb_obj: the current Excel workbook
    :param current_scale: str - current scale to skip it
    :param codes: list - relationship codes
    :param start_row: int - row number where the first query is located
    :param scales_column: int - column number where the scales are located
    :param related_column: int - column number where the relationship between queries in different scales is located
    :return: result: list - the list with the scale names

    """

    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row

    result = []

    for code in codes:

        for i in range(m_row, start_row - 1, -1):
            related_code = str(sheet_obj.cell(row=i, column=related_column).value)

            if related_code != "None":  # Empty cell

                codes = convert_code(related_code)

                if code in codes:  # Code is present in cell
                    scale = sheet_obj.cell(row=i, column=scales_column).value

                    # Row with code, but maybe it doesn't have the scale name
                    # (merged cells - go back until we find the name)
                    j = i
                    while scale is None:
                        scale = sheet_obj.cell(row=j, column=scales_column).value
                        j -= 1

                    if scale not in result and scale != current_scale:
                        result.append(scale)

    return result


def convert_code(code: str):
    """
    Function that given a query relationship code returns a list with the desired code format

    :param code: str - the code to format
    :return: result: list - the list of codes
    """
    result = []
    aux = []

    if "_" in code:
        aux = code.split("_")
    if len(aux) > 0:
        for c in aux:
            result.append(c)
    else:
        result.append(code)

    return result


def code_present(wb_obj, code: str, start_row: int, related_column: int):
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row

    found = False  # scale associated has been found
    for i in range(start_row, m_row + 1):
        related_code = str(sheet_obj.cell(row=i, column=related_column).value)
        if str(related_code) == code:
            found = True
    return found


queries_and_scales = get_queries_and_scales("./excel/scales.xlsx", 5, 2, 4, 5)
print(queries_and_scales)
